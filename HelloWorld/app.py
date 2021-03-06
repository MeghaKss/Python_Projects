import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook(filename="transactions.xlsx")
sheet = wb['Sheet1']
# cell = sheet['A1']
# print(sheet.max_row)
for row in range(2, sheet.max_row + 1):
    # print(row)
    cell = sheet.cell(row=row, column=3)
    # changes here -specified the argument
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row=row, column=4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(data=values)
sheet.add_chart(chart, 'e2')

wb.save(filename="transactions2.xlsx")

